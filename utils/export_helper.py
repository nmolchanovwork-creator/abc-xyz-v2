"""Модуль экспорта результатов: сводная таблица и история AI-стратегий в Excel."""

import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def summary_to_excel_bytes(summary_df: pd.DataFrame, filters_desc: str = "") -> bytes:
    """Формирует Excel-файл со сводной таблицей ABC-XYZ анализа для скачивания."""
    export_cols = [
        "SKU", "Наименование", "Категория", "Продажи шт.", "Выручка", "Маржа",
        "Текущий_остаток", "ABC_код", "XYZ_код", "Итоговый_статус", "AI Совет",
    ]
    export_cols = [c for c in export_cols if c in summary_df.columns]
    df = summary_df[export_cols].copy()

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ABC-XYZ анализ", startrow=2 if filters_desc else 0)
        ws = writer.sheets["ABC-XYZ анализ"]

        header_row = 3 if filters_desc else 1
        if filters_desc:
            ws.cell(row=1, column=1, value=f"Фильтры: {filters_desc}")

        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(export_cols) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_idx, col_name in enumerate(export_cols, start=1):
            letter = get_column_letter(col_idx)
            width = max(14, min(40, len(col_name) + 6))
            ws.column_dimensions[letter].width = width

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    return buffer.getvalue()


def strategy_history_to_text(history: list) -> str:
    """Собирает всю историю AI-стратегий в один текстовый файл для скачивания."""
    parts = []
    for i, entry in enumerate(history, start=1):
        parts.append(
            f"=== Стратегия №{i} | {entry['timestamp']} | "
            f"Провайдер: {entry['provider']} / {entry['model']} ===\n"
            f"Фильтры: {entry['filters_desc']}\n\n"
            f"{entry['text']}\n"
        )
    return "\n\n".join(parts)
