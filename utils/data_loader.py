"""
Модуль загрузки данных.

Единственный источник данных — файл, загруженный пользователем (CSV / Excel).
Никаких внешних подключений (Google Sheets и т.п.) не требуется.

Также содержит:
- генератор шаблона Excel-файла для скачивания (с примером данных и листом-инструкцией);
- генератор демо-данных — чтобы можно было опробовать приложение без своего файла.
"""

import io
import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MONTH_ORDER = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
               "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]

# Колонки, которые ОБЯЗАТЕЛЬНО должны быть в файле пользователя
BASE_REQUIRED_COLUMNS = [
    "SKU", "Наименование", "Категория", "Тип продажи",
    "Месяц", "Год", "Себестоимость", "Цена",
    "Продажи шт.", "Текущий остаток",
]

# Эти колонки не обязательны — если их нет (или есть пустые значения),
# приложение посчитает их автоматически: Выручка = Продажи*Цена,
# Маржа = Продажи*(Цена-Себестоимость)
DERIVED_COLUMNS = ["Выручка", "Маржа"]

ALL_TEMPLATE_COLUMNS = BASE_REQUIRED_COLUMNS[:8] + ["Продажи шт.", "Выручка", "Маржа", "Текущий остаток"]


class DataValidationError(Exception):
    """Исключение для ошибок валидации структуры/содержимого данных."""
    pass


def validate_columns(df: pd.DataFrame) -> None:
    """Проверяет, что в таблице присутствуют все обязательные (не вычисляемые) колонки."""
    missing = [c for c in BASE_REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise DataValidationError(
            f"В загруженном файле отсутствуют обязательные колонки: {', '.join(missing)}. "
            f"Скачайте шаблон в боковой панели, чтобы увидеть правильную структуру."
        )


def _post_process(df: pd.DataFrame) -> pd.DataFrame:
    """Приводит типы, чистит пустые SKU, автоматически считает Выручку/Маржу если их нет."""
    validate_columns(df)

    numeric_base = ["Себестоимость", "Цена", "Продажи шт.", "Текущий остаток", "Год"]
    for col in numeric_base:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    df = df.dropna(subset=["SKU"] + numeric_base)
    if df.empty:
        raise DataValidationError("После обработки данных не осталось ни одной корректной строки. "
                                   "Проверьте числовые колонки (Себестоимость, Цена, Продажи шт., "
                                   "Текущий остаток, Год) — там не должно быть текста или пустых ячеек.")

    # Выручка: если колонки нет или есть пропуски — считаем сами
    if "Выручка" not in df.columns:
        df["Выручка"] = df["Продажи шт."] * df["Цена"]
    else:
        df["Выручка"] = pd.to_numeric(df["Выручка"], errors="coerce")
        calc = df["Продажи шт."] * df["Цена"]
        df["Выручка"] = df["Выручка"].fillna(calc)

    # Маржа: если колонки нет или есть пропуски — считаем сами
    if "Маржа" not in df.columns:
        df["Маржа"] = df["Продажи шт."] * (df["Цена"] - df["Себестоимость"])
    else:
        df["Маржа"] = pd.to_numeric(df["Маржа"], errors="coerce")
        calc = df["Продажи шт."] * (df["Цена"] - df["Себестоимость"])
        df["Маржа"] = df["Маржа"].fillna(calc)

    df["Месяц"] = df["Месяц"].astype(str).str.strip().str.capitalize()
    df["Категория"] = df["Категория"].astype(str).str.strip()
    df["Тип продажи"] = df["Тип продажи"].astype(str).str.strip()

    return df


@st.cache_data(show_spinner="Обработка загруженного файла...")
def load_from_upload(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Загружает данные из CSV или Excel файла."""
    try:
        if filename.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_bytes))
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)
    except Exception as e:
        raise DataValidationError(f"Не удалось прочитать файл: {e}")
    return _post_process(df)


@st.cache_data(show_spinner=False)
def generate_demo_data(n_sku: int = 60, months: int = 12, seed: int = 42) -> pd.DataFrame:
    """Генерирует демонстрационный набор данных (чтобы опробовать приложение без своего файла)."""
    rng = np.random.default_rng(seed)
    categories = ["Электроника", "Одежда", "Дом", "Спорт", "Красота"]
    sale_types = ["розница", "опт", "индивидуально"]

    rows = []
    for i in range(n_sku):
        sku = f"SKU-{1000 + i}"
        base_qty = rng.integers(5, 500)
        cost = round(float(rng.uniform(50, 5000)), 2)
        price = round(cost * rng.uniform(1.1, 2.5), 2)
        cat = categories[i % len(categories)]
        stype = sale_types[i % len(sale_types)]
        volatility = rng.uniform(0.05, 0.6)
        for m in range(months):
            qty = max(0, int(rng.normal(base_qty, base_qty * volatility)))
            revenue = round(qty * price, 2)
            margin = round(qty * (price - cost), 2)
            rows.append({
                "SKU": sku,
                "Наименование": f"Товар {i + 1}",
                "Категория": cat,
                "Тип продажи": stype,
                "Месяц": MONTH_ORDER[m % 12],
                "Год": 2025,
                "Себестоимость": cost,
                "Цена": price,
                "Продажи шт.": qty,
                "Выручка": revenue,
                "Маржа": margin,
                "Текущий остаток": int(rng.integers(0, base_qty * 2)),
            })
    return pd.DataFrame(rows)


def generate_template_excel() -> bytes:
    """
    Формирует Excel-шаблон для скачивания:
    - лист «Данные» — заголовки + 3 примера строк;
    - лист «Инструкция» — описание каждой колонки, формат, обязательность.
    """
    wb = Workbook()

    # --- Лист "Данные" ---
    ws = wb.active
    ws.title = "Данные"

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(ALL_TEMPLATE_COLUMNS)
    for col_idx in range(1, len(ALL_TEMPLATE_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    examples = [
        ["SKU-001", "Кроссовки беговые Air", "Обувь", "розница", "Январь", 2025, 1500, 3200, 45, "", "", 60],
        ["SKU-001", "Кроссовки беговые Air", "Обувь", "розница", "Февраль", 2025, 1500, 3200, 38, "", "", 60],
        ["SKU-002", "Куртка зимняя", "Одежда", "опт", "Январь", 2025, 4200, 8900, 12, "", "", 20],
    ]
    for row in examples:
        ws.append(row)

    for col_idx, col_name in enumerate(ALL_TEMPLATE_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = max(14, len(col_name) + 4)
    ws.freeze_panes = "A2"

    # --- Лист "Инструкция" ---
    ws2 = wb.create_sheet("Инструкция")
    ws2.append(["Колонка", "Обязательна?", "Формат / описание"])
    for col_idx in range(1, 4):
        cell = ws2.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    instructions = [
        ["SKU", "Да", "Уникальный код товара, повторяется для каждой строки-месяца"],
        ["Наименование", "Да", "Название товара"],
        ["Категория", "Да", "Категория товара (используется как фильтр)"],
        ["Тип продажи", "Да", "розница / опт / индивидуально (используется как фильтр)"],
        ["Месяц", "Да", "Название месяца по-русски: Январь, Февраль, ... Декабрь"],
        ["Год", "Да", "Год числом, например 2025"],
        ["Себестоимость", "Да", "Себестоимость единицы товара, число без валютных символов"],
        ["Цена", "Да", "Цена продажи единицы товара, число без валютных символов"],
        ["Продажи шт.", "Да", "Количество проданных единиц за месяц"],
        ["Выручка", "Нет", "Если оставить пустым — будет посчитано как Продажи шт. × Цена"],
        ["Маржа", "Нет", "Если оставить пустым — будет посчитано как Продажи шт. × (Цена − Себестоимость)"],
        ["Текущий остаток", "Да", "Остаток на складе на текущий момент, шт."],
    ]
    for row in instructions:
        ws2.append(row)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 75
    for r in ws2.iter_rows(min_row=2, max_row=len(instructions) + 1):
        for cell in r:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws2.append([])
    ws2.append(["Важно: каждая строка = один SKU за один месяц одного года.",
                "", "Если у товара есть продажи в нескольких месяцах — для него должно быть "
                    "несколько строк (по одной на месяц). Это нужно для расчёта XYZ-анализа."])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
